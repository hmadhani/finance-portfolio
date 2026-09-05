#!/usr/bin/env python3
"""
Portfolio XLS Report Generator — finance-portfolio (isolated, market-data-only)

Generates a quarterly or annual performance workbook from portfolio.md's
current state, framed as prepared for "Investor: Mr. Spock — Logical
Investor" (see investor-profile.md). Self-contained: no dependency on any
file outside this repo, since the cloud routine that runs this only has
finance-portfolio checked out.
"""

import argparse
from typing import Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, PieChart, Reference
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl --break-system-packages")
    exit(1)


# Style definitions (ported from financial-advisor-skill's excel_report_generator.py —
# verified generic, no personal data, no vault references)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PERCENT_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER


def apply_subheader_style(cell):
    """Apply subheader styling to a cell."""
    cell.font = Font(bold=True)
    cell.fill = SUBHEADER_FILL
    cell.border = THIN_BORDER


def auto_column_width(ws, min_width=10, max_width=50):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_cover_sheet(wb: Workbook, portfolio_data: Dict, period_label: str) -> None:
    """Cover sheet framing the report as prepared for the Mr. Spock investor profile."""
    ws = wb.create_sheet("Cover", 0)

    ws['A1'] = "PORTFOLIO PERFORMANCE REPORT"
    ws['A1'].font = Font(bold=True, size=18, color="2F5496")
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Period: {period_label}"
    ws['A2'].font = Font(italic=True, size=11)

    ws['A4'] = "Prepared for: Mr. Spock — Logical Investor"
    ws['A4'].font = Font(bold=True, size=13)

    ws['A5'] = (
        "A hypothetical growth-oriented investor in the accumulation phase, "
        "governed by a fixed sizing table, a 25% theme-concentration cap, and "
        "a wall-clock-derived glide path. See investor-profile.md for the full "
        "mandate. Zero real personal data underlies this profile."
    )
    ws['A5'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A5:D8')

    row = 10
    ws.cell(row=row, column=1, value="Glide-Path Phase").font = Font(bold=True)
    ws.cell(row=row, column=2, value=portfolio_data.get('glide_path_phase', 'N/A'))
    row += 1
    ws.cell(row=row, column=1, value="Cycle #").font = Font(bold=True)
    ws.cell(row=row, column=2, value=portfolio_data.get('cycle_number', 'N/A'))
    row += 2

    ws.cell(row=row, column=1, value=(
        "IMPORTANT: This is a paper/model portfolio for evaluation purposes "
        "only, tracked with public market data alone (no personal financial "
        "context). It does not constitute investment advice or a "
        "recommendation to buy or sell any security. All investments involve "
        "risk, including potential loss of principal. Always consult with "
        "qualified financial professionals before making investment decisions."
    ))
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(f'A{row}:D{row+3}')

    auto_column_width(ws)


def create_summary_sheet(wb: Workbook, portfolio_data: Dict, benchmark_data: Dict) -> None:
    """NAV, return vs. blended benchmark vs. SPY, sleeve drift."""
    ws = wb.create_sheet("Summary")

    ws['A1'] = "SUMMARY"
    ws['A1'].font = Font(bold=True, size=14, color="2F5496")

    headers = ["Metric", "This Portfolio", "Blended Benchmark", "SPY"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        apply_header_style(cell)

    rows = [
        ("NAV", portfolio_data.get('nav'), None, None),
        ("Period Return %", portfolio_data.get('period_return_pct'),
         benchmark_data.get('blended_return_pct'), benchmark_data.get('spy_return_pct')),
        ("Since-Inception Return %", portfolio_data.get('inception_return_pct'),
         benchmark_data.get('blended_inception_return_pct'), benchmark_data.get('spy_inception_return_pct')),
    ]
    r = 4
    for label, this_val, bench_val, spy_val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=this_val)
        ws.cell(row=r, column=3, value=bench_val)
        ws.cell(row=r, column=4, value=spy_val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="SLEEVE DRIFT VS. TARGET").font = Font(bold=True, size=12, color="2F5496")
    r += 1
    headers2 = ["Sleeve", "Target %", "Current %", "Drift"]
    for col, h in enumerate(headers2, start=1):
        cell = ws.cell(row=r, column=col, value=h)
        apply_subheader_style(cell)
    r += 1
    for sleeve in portfolio_data.get('sleeve_drift', []):
        ws.cell(row=r, column=1, value=sleeve['name'])
        ws.cell(row=r, column=2, value=sleeve['target_pct'])
        ws.cell(row=r, column=3, value=sleeve['current_pct'])
        ws.cell(row=r, column=4, value=sleeve['current_pct'] - sleeve['target_pct'])
        r += 1

    auto_column_width(ws)


def create_holdings_sheet(wb: Workbook, holdings: List[Dict]) -> None:
    """Current positions + theme tags."""
    ws = wb.create_sheet("Holdings")
    headers = ["Ticker", "Type", "Theme", "Shares", "Entry Price", "Current Price", "Mkt Value", "% NAV", "Unrealized G/L"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    for r, h in enumerate(holdings, start=2):
        ws.cell(row=r, column=1, value=h['ticker'])
        ws.cell(row=r, column=2, value=h['type'])
        ws.cell(row=r, column=3, value=h['theme'])
        ws.cell(row=r, column=4, value=h.get('shares'))
        ws.cell(row=r, column=5, value=h.get('entry_price'))
        ws.cell(row=r, column=6, value=h.get('current_price'))
        ws.cell(row=r, column=7, value=h['mkt_value'])
        ws.cell(row=r, column=8, value=h['pct_nav'])
        ws.cell(row=r, column=9, value=h.get('unrealized_gl'))

    auto_column_width(ws)


def create_trade_log_sheet(wb: Workbook, trades: List[Dict]) -> None:
    """Period's trades with rationale."""
    ws = wb.create_sheet("Trade Log")
    headers = ["Date", "Ticker", "Action", "Shares", "Price", "Rationale"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    for r, t in enumerate(trades, start=2):
        ws.cell(row=r, column=1, value=t['date'])
        ws.cell(row=r, column=2, value=t['ticker'])
        ws.cell(row=r, column=3, value=t['action'])
        ws.cell(row=r, column=4, value=t.get('shares'))
        ws.cell(row=r, column=5, value=t.get('price'))
        cell = ws.cell(row=r, column=6, value=t['rationale'])
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions['F'].width = 80
    auto_column_width(ws, max_width=20)


def create_charts_sheet(wb: Workbook, theme_exposure: List[Dict], return_trend: List[Dict]) -> None:
    """Allocation pie, sector/theme exposure, return trend — mirrors portfolio-dashboard.html."""
    ws = wb.create_sheet("Charts")

    ws['A1'] = "Theme"
    ws['B1'] = "% NAV"
    for r, t in enumerate(theme_exposure, start=2):
        ws.cell(row=r, column=1, value=t['theme'])
        ws.cell(row=r, column=2, value=t['pct_nav'])

    pie = PieChart()
    pie.title = "Theme Exposure"
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(theme_exposure))
    labels = Reference(ws, min_col=1, min_row=2, max_row=1 + len(theme_exposure))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "D1")

    trend_start_row = len(theme_exposure) + 4
    ws.cell(row=trend_start_row, column=1, value="Cycle Date")
    ws.cell(row=trend_start_row, column=2, value="NAV")
    ws.cell(row=trend_start_row, column=3, value="Blended Benchmark")
    ws.cell(row=trend_start_row, column=4, value="SPY")
    for r, point in enumerate(return_trend, start=trend_start_row + 1):
        ws.cell(row=r, column=1, value=point['date'])
        ws.cell(row=r, column=2, value=point['nav'])
        ws.cell(row=r, column=3, value=point['blended_benchmark'])
        ws.cell(row=r, column=4, value=point['spy'])

    line = LineChart()
    line.title = "Return Trend"
    data = Reference(ws, min_col=2, max_col=4, min_row=trend_start_row, max_row=trend_start_row + len(return_trend))
    cats = Reference(ws, min_col=1, min_row=trend_start_row + 1, max_row=trend_start_row + len(return_trend))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, f"D{trend_start_row}")

    auto_column_width(ws)


def create_portfolio_performance_report(
    portfolio_data: Dict[str, Any],
    benchmark_data: Dict[str, Any],
    period_label: str,
    output_path: str,
) -> str:
    """Build the full quarterly/annual workbook. Returns output_path."""
    wb = Workbook()
    wb.remove(wb.active)

    create_cover_sheet(wb, portfolio_data, period_label)
    create_summary_sheet(wb, portfolio_data, benchmark_data)
    create_holdings_sheet(wb, portfolio_data['holdings'])
    create_trade_log_sheet(wb, portfolio_data['trades'])
    create_charts_sheet(wb, portfolio_data['theme_exposure'], portfolio_data['return_trend'])

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate finance-portfolio XLS performance report")
    parser.add_argument("--period", required=True, help="Period label, e.g. 2026-Q3 or 2026-annual")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--data-json", required=True, help="Path to a JSON file with portfolio_data + benchmark_data (produced by the routine from portfolio.md)")
    args = parser.parse_args()

    import json
    with open(args.data_json) as f:
        payload = json.load(f)

    path = create_portfolio_performance_report(
        payload['portfolio_data'], payload['benchmark_data'], args.period, args.output
    )
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
